from openpyxl import load_workbook
from colorama import Fore
from configs import tg
import telebot
from transliterate import translit

bot = telebot.TeleBot(tg.TOKEN)


def generate(message, paht_file):
    try:
        print(Fore.BLUE + f'Получил файл {paht_file}')
        # system_name_diitionary = "Doczilla_bot"
        system_name_dictionary1 = f"{translit(paht_file.split('/')[-1].split('.')[0],language_code='ru',reversed=True)}"
        system_name_dictionary = ''
        for name_dict in list(system_name_dictionary1):
            if name_dict == "'":
                print(name_dict)
                name_dict = ''
                system_name_dictionary += name_dict
            system_name_dictionary += name_dict
        wb = load_workbook(f'{paht_file}')
        ws = wb[wb.sheetnames[0]]
        dict_name = system_name_dictionary
        imports = []
        controls = []
        save_name = paht_file.split('/')[-1].split('.')[0]
        # save_name = "doczilla_bot"
        lines = ["import org.zenframework.z8.base.table.Table;"
            , "import pro.doczilla.dictionary.importFile.CsvDictionaryImportAction;", " "
            , "[generatable]", "[dictionary]", "[request true]", f'[name "{dict_name}"]',
                 f'[displayName "{save_name}"]',
                 f"public class {dict_name} extends Table", "{"]

        for cell in ws[2]:  # ws[номер строчки с данными]
            if str(type(cell.value)) == "<class 'str'>":
                imports.append('import org.zenframework.z8.base.table.value.StringField;')
                imports.append("//Справочник сделан при поддержке телеграм бота https://t.me/Doczilla_bot\n")
            elif str(type(cell.value)) == "<class 'int'>":
                imports.append('import org.zenframework.z8.base.table.value.DecimalField;')
            elif str(type(cell.value)) == "<class 'datetime.datetime'>":
                imports.append('import org.zenframework.z8.base.table.value.DateField;')
            elif str(type(cell.value)) == "<class 'float'>":
                imports.append('import org.zenframework.z8.base.table.value.DecimalField;')
            else:
                print(type(cell.value))
                imports.append('import org.zenframework.z8.base.table.value.StringField;')
                imports.append("//Справочник сделан при поддержке телеграм бота https://t.me/Doczilla_bot\n")

        with open(rf"temp/bl/{save_name}.txt", "w") as file:
            for line2 in set(imports):
                file.write(line2 + '\n')
            for line in lines:
                file.write(line + '\n')

        with open(f'temp/bl/{save_name}.txt', 'r') as f:
            lines = f.read()

        with open(f'temp/bl/{save_name}.txt', 'w') as newf:
            newf.write(lines)
            i = len(ws[2])
            c = -1
            while i >= 1:
                i -= 1
                c += 1
                text = str(ws[1][c].value)
                name_column1 = f"{translit(text, language_code='ru', reversed=True)}"
                name_column = ''
                for sim in list(name_column1):
                    if sim == "'":
                        print(name_column)
                        sim = ''
                        name_column += sim
                    name_column += sim
                display_name_column = ws[1][c].value
                types = type(ws[2][c].value)
                if str(types) == "<class 'str'>":
                    tip = 'StringField'
                elif str(types) == "<class 'int'>":
                    tip = 'DecimalField'
                elif str(types) == "<class 'datetime.datetime'>":
                    tip = 'DateField'
                elif str(types) == "<class 'float'>":
                    tip = 'DecimalField'
                else:
                    tip = 'StringField'
                if tip == 'DateField':
                    column_list = [f'  [name "{name_column}"]', f'  [displayName "{display_name_column}"]',
                                   f'  public {tip} {name_column};', f'  {name_column}.colSpan = 1;\n']
                    controls.append(name_column)
                else:
                    column_list = [f'  [name "{name_column}"]', f'  [displayName "{display_name_column}"]',
                                   f'  public {tip} {name_column};', f'  {name_column}.length = 200;',
                                   f'  {name_column}.colSpan = 1;\n']
                    controls.append(name_column)
                for line in column_list:
                    newf.write(line + f"\n")
            else:
                print('Справочник готов', save_name)

        with open(f'temp/bl/{save_name}.txt', 'r') as f:
            lines2 = f.read()

        with open(rf"temp/bl/{save_name}.txt", "w", encoding="utf-8") as file:
            file.write(lines2 + '\n')
            m = ','.join(map(str, controls))
            file.write('  columns = {' + m + '};' + '\n')
            file.write('  controls  = {' + m + '};' + '\n')
            file.write('  actions = {importAction};' + '\n')
            file.write('  CsvDictionaryImportAction importAction;' + '\n')
            file.write('}' + '\n')

        doc = open(f'temp/bl/{save_name}.txt', 'rb')
        bot.send_document(message.chat.id, doc)
    except  Exception as e:
        bot.send_message(message.chat.id,
                         f'ОШИБКА!\n{e}')
